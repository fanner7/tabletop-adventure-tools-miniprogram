# -*- coding: utf-8 -*-
"""Check if 人物卡 has 备注 column specifically."""
import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('COC7空白卡CY2lusFinal (1).xlsx', data_only=True)
ws = wb.worksheets[2]  # 人物卡

# Check skill rows (16-49) for col 5, 7, 9, 11, 13, 15 specifically
print("=== 人物卡 skill rows with col 5,7,9,11,13,15 ===")
for r in range(16, 50):
    name = ws.cell(row=r, column=6).value
    if not name or not str(name).strip():
        continue
    n = str(name).strip()
    
    # Check columns 5, 7, 9, 11, 13, 15
    extras = {}
    for c in [5, 7, 9, 11, 13, 15]:
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            extras[c] = str(v).strip()[:200]
    if extras:
        print(f"Row {r} [{n}]: {json.dumps(extras, ensure_ascii=False)}")

# Also check the header row (row 15) for column labels
print("\n=== Row 15 (header) all columns ===")
for c in range(1, 90):
    v = ws.cell(row=15, column=c).value
    if v is not None and str(v).strip():
        print(f"  col {c}: {str(v).strip()[:80]}")
