# -*- coding: utf-8 -*-
"""Check 附表 sheet structure and extract skill descriptions properly."""
import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('COC7空白卡CY2lusFinal (1).xlsx', data_only=True)
ws = wb.worksheets[7]  # 附表

# Check header rows
print("=== 附表 headers (rows 1-5) ===")
for r in range(1, 6):
    row_data = {}
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            row_data[c] = str(v).strip()[:100]
    print(f"Row {r}: {json.dumps(row_data, ensure_ascii=False)}")

# Check what columns are used around row 95+
print("\n=== 附表 columns for row 95+ ===")
columns_used = set()
for r in range(95, 213):
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            columns_used.add(c)
print(f"Columns used: {sorted(columns_used)}")

# Print first few data rows
print("\n=== Sample data rows ===")
for r in range(95, 110):
    b = ws.cell(row=r, column=2).value  # name
    e = ws.cell(row=r, column=5).value  # 特殊适用
    f = ws.cell(row=r, column=6).value  # desc
    g = ws.cell(row=r, column=7).value
    h = ws.cell(row=r, column=8).value
    i = ws.cell(row=r, column=9).value
    j = ws.cell(row=r, column=10).value
    k = ws.cell(row=r, column=11).value
    
    if b and str(b).strip():
        parts = [f"name={str(b).strip()[:30]}"]
        if e: parts.append(f"usage={str(e).strip()[:50]}")
        if f: parts.append(f"desc={str(f).strip()[:80]}")
        if g: parts.append(f"col7={str(g).strip()[:50]}")
        if h: parts.append(f"col8={str(h).strip()[:50]}")
        if i: parts.append(f"col9={str(i).strip()[:50]}")
        if j: parts.append(f"col10={str(j).strip()[:50]}")
        if k: parts.append(f"col11={str(k).strip()[:50]}")
        print(f"Row {r}: {' | '.join(parts)}")
