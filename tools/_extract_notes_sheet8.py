# -*- coding: utf-8 -*-
"""Deep-extract from 技能注释 sheet (sheet 8) — this may contain skill click notes."""
import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('COC7空白卡CY2lusFinal (1).xlsx', data_only=True)
ws = wb.worksheets[8]  # 技能注释

print(f"技能注释: rows={ws.max_row}, cols={ws.max_column}")

# Print ALL rows comprehensively
for r in range(1, ws.max_row + 1):
    row_data = {}
    for c in range(1, min(ws.max_column + 1, 30)):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            row_data[c] = str(v).strip()[:200]
    if row_data:
        print(f"\nRow {r}:")
        for c, val in row_data.items():
            print(f"  col {c}: {val}")
