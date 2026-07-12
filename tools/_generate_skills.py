# -*- coding: utf-8 -*-
"""Extract all skill data from Excel 附表 and generate valid JavaScript for ALL_SKILLS.
Outputs to a preview file first for verification."""
import openpyxl, json, re, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('COC7空白卡CY2lusFinal (1).xlsx', data_only=True)
ws = wb.worksheets[7]  # 附表

# ============================================================
# 1. Extract skill data from 附表
# ============================================================
# Columns:
#   B(2): 技能名称
#   E(5): 特殊适用
#   F(6): 技能解释 (desc)
#   G(7): 普通难度
#   H(8): 困难难度
#   I(9): 孤注一掷示例
#   J(10): 孤注一掷失败的范例结果
#   K(11): 疯狂调查员孤注一掷失败的范例结果

raw_skills = {}
for r in range(95, 213):
    name = ws.cell(row=r, column=2).value
    if not name or not str(name).strip():
        continue
    n = str(name).strip()
    if n in ('——————————',):
        continue
    
    def cell_str(row, col):
        v = ws.cell(row=row, column=col).value
        return str(v).strip() if v and str(v).strip() else ''
    
    raw_skills[n] = {
        'name': n,
        'usage': cell_str(r, 5),      # 特殊适用
        'desc': cell_str(r, 6),       # 技能解释
        'diff_normal': cell_str(r, 7),  # 普通难度
        'diff_hard': cell_str(r, 8),    # 困难难度
        'push_example': cell_str(r, 9), # 孤注一掷示例
        'push_fail': cell_str(r, 10),   # 孤注一掷失败的范例结果
        'push_insane': cell_str(r, 11), # 疯狂调查员孤注一掷失败的范例结果
    }

print(f"Extracted {len(raw_skills)} skills from 附表")

# ============================================================
# 2. Skill definitions (matching current ALL_SKILLS structure)
# ============================================================
# These are the skills that appear in ALL_SKILLS with their base/cat/related
skills_def = [
    ("会计", 5, "investigate", "INT"),
    ("人类学", 1, "investigate", "INT"),
    ("估价", 5, "investigate", "INT"),
    ("考古学", 1, "investigate", "INT"),
    ("侦查", 25, "investigate", "INT"),
    ("聆听", 20, "investigate", "INT"),
    ("追踪", 10, "investigate", "INT"),
    ("图书馆使用", 20, "investigate", "INT"),
    ("计算机使用", 5, "investigate", "INT"),
    ("信用评级", 0, "investigate", "INT"),
    ("克苏鲁神话", 0, "investigate", "INT"),
    ("锁匠", 1, "investigate", "DEX"),
    ("妙手", 10, "investigate", "DEX"),
    ("导航", 10, "investigate", "INT"),
    ("话术", 5, "social", "APP"),
    ("说服", 10, "social", "APP"),
    ("恐吓", 15, "social", "STR"),
    ("取悦", 15, "social", "APP"),
    ("乔装", 5, "social", "APP"),
    ("心理学", 10, "social", "INT"),
    ("投掷", 20, "combat", "DEX"),
    ("闪避", 0, "combat", "DEX/2"),
    ("攀爬", 20, "special", "STR"),
    ("跳跃", 20, "special", "STR"),
    ("游泳", 20, "special", "STR"),
    ("潜行", 20, "special", "DEX"),
    ("骑术", 5, "special", "DEX"),
    ("急救", 30, "support", "INT"),
    ("医学", 1, "support", "INT"),
    ("精神分析", 1, "support", "INT"),
    ("电气维修", 10, "knowledge", "INT"),
    ("电子学", 1, "knowledge", "INT"),
    ("法律", 5, "knowledge", "INT"),
    ("历史", 5, "knowledge", "INT"),
    ("母语", 0, "knowledge", "EDU"),
    ("机械维修", 10, "knowledge", "INT"),
    ("博物学", 10, "knowledge", "INT"),
    ("神秘学", 5, "knowledge", "INT"),
    ("科学", 1, "knowledge", "INT"),
    ("操作重型机械", 1, "knowledge", "INT"),
    ("药学", 1, "knowledge", "INT"),
    ("催眠", 1, "knowledge", "INT"),
    ("读唇", 1, "knowledge", "INT"),
    ("爆破", 1, "knowledge", "INT"),
    ("潜水", 1, "knowledge", "INT"),
    ("动物驯养", 5, "knowledge", "INT"),
]

# Skills with spec (no desc, or both)
spec_skills = {
    "格斗①": {"base": 25, "cat": "combat", "related": "STR/DEX",
               "spec": '{"type":"options","defaultBase":25,"options":[{"name":"斗殴","base":25},{"name":"鞭子","base":5},{"name":"电锯","base":10},{"name":"链枷","base":10},{"name":"绞具","base":15},{"name":"斧","base":15},{"name":"剑","base":20},{"name":"矛","base":20}]}'},
    "格斗②": {"base": 25, "cat": "combat", "related": "STR/DEX",
               "spec": '{"type":"options","defaultBase":25,"options":[{"name":"斗殴","base":25},{"name":"鞭子","base":5},{"name":"电锯","base":10},{"name":"链枷","base":10},{"name":"绞具","base":15},{"name":"斧","base":15},{"name":"剑","base":20},{"name":"矛","base":20}]}'},
    "格斗③": {"base": 25, "cat": "combat", "related": "STR/DEX",
               "spec": '{"type":"text","defaultBase":25}'},
    "射击①": {"base": 20, "cat": "combat", "related": "DEX",
               "spec": '{"type":"options","defaultBase":20,"options":[{"name":"手枪","base":20},{"name":"步枪/霰弹枪","base":25},{"name":"冲锋枪","base":15},{"name":"机枪","base":10},{"name":"弓术","base":15},{"name":"喷射器","base":10},{"name":"重武器","base":10}]}'},
    "射击②": {"base": 20, "cat": "combat", "related": "DEX",
               "spec": '{"type":"options","defaultBase":20,"options":[{"name":"手枪","base":20},{"name":"步枪/霰弹枪","base":25},{"name":"冲锋枪","base":15},{"name":"机枪","base":10},{"name":"弓术","base":15},{"name":"喷射器","base":10},{"name":"重武器","base":10}]}'},
    "射击③": {"base": 15, "cat": "combat", "related": "DEX",
               "spec": '{"type":"text","defaultBase":15}'},
    "技艺①": {"base": 5, "cat": "knowledge", "related": "INT",
               "spec": '{"type":"options","defaultBase":5,"options":[{"name":"表演","base":5},{"name":"美术","base":5},{"name":"摄影","base":5},{"name":"伪造","base":5},{"name":"写作","base":5},{"name":"书法","base":5},{"name":"乐理","base":5},{"name":"厨艺","base":5},{"name":"裁缝","base":5},{"name":"理发","base":5},{"name":"建筑","base":5},{"name":"舞蹈","base":5},{"name":"酿酒","base":5},{"name":"捕鱼","base":5},{"name":"歌唱","base":5},{"name":"制陶","base":5},{"name":"雕塑","base":5},{"name":"杂技","base":5},{"name":"风水","base":5},{"name":"技术制图","base":5},{"name":"耕作","base":5},{"name":"打字","base":5},{"name":"速记","base":5},{"name":"木匠","base":5},{"name":"莫里斯舞蹈","base":5},{"name":"歌剧歌唱","base":5},{"name":"粉刷匠与油漆工","base":5},{"name":"吹真空管","base":5}]}'},
    "技艺②": {"base": 5, "cat": "knowledge", "related": "INT",
               "spec": '{"type":"options","defaultBase":5,"options":[{"name":"表演","base":5},{"name":"美术","base":5},{"name":"摄影","base":5},{"name":"伪造","base":5},{"name":"写作","base":5},{"name":"书法","base":5},{"name":"乐理","base":5},{"name":"厨艺","base":5},{"name":"裁缝","base":5},{"name":"理发","base":5},{"name":"建筑","base":5},{"name":"舞蹈","base":5},{"name":"酿酒","base":5},{"name":"捕鱼","base":5},{"name":"歌唱","base":5},{"name":"制陶","base":5},{"name":"雕塑","base":5},{"name":"杂技","base":5},{"name":"风水","base":5},{"name":"技术制图","base":5},{"name":"耕作","base":5},{"name":"打字","base":5},{"name":"速记","base":5},{"name":"木匠","base":5},{"name":"莫里斯舞蹈","base":5},{"name":"歌剧歌唱","base":5},{"name":"粉刷匠与油漆工","base":5},{"name":"吹真空管","base":5}]}'},
    "技艺③": {"base": 5, "cat": "knowledge", "related": "INT",
               "spec": '{"type":"text","defaultBase":5}'},
    "外语①": {"base": 1, "cat": "knowledge", "related": "INT",
               "spec": '{"type":"text","defaultBase":1}'},
    "外语②": {"base": 1, "cat": "knowledge", "related": "INT",
               "spec": '{"type":"text","defaultBase":1}'},
    "外语③": {"base": 1, "cat": "knowledge", "related": "INT",
               "spec": '{"type":"text","defaultBase":1}'},
    "驾驶①": {"base": 20, "cat": "special", "related": "DEX",
               "spec": '{"type":"text","defaultBase":20}'},
    "科学": {"base": 1, "cat": "knowledge", "related": "INT",
             "spec": '{"type":"options","defaultBase":1,"options":[{"name":"数学","base":10},{"name":"地质学","base":1},{"name":"化学","base":1},{"name":"生物学","base":1},{"name":"物理学","base":1},{"name":"天文学","base":1},{"name":"气象学","base":1},{"name":"药学","base":1},{"name":"工程学","base":1},{"name":"密码学","base":1},{"name":"制图学","base":1},{"name":"人类学","base":1},{"name":"心理学","base":1}]}'},
}

# ============================================================
# 3. Escape function for JS single-quoted strings
# ============================================================
def js_escape(text):
    """Escape text for use inside a JS single-quoted string."""
    if not text:
        return ''
    # Order matters: backslash first, then newlines, then quotes
    text = text.replace('\\', '\\\\')  # escape backslashes
    text = text.replace('\n', '\\n')   # escape newlines
    text = text.replace("'", "\\'")    # escape single quotes
    return text

# ============================================================
# 4. Generate JS lines
# ============================================================
lines = []
lines.append('const ALL_SKILLS = [')

# Helper: Excel name → skill name mapping
excel_name_map = {
    "计算机使用 Ω": "计算机使用",
    "电子学 Ω": "电子学",
}

def get_excel_data(name):
    """Try to find Excel data by name, with fallbacks."""
    if name in raw_skills:
        return raw_skills[name]
    # Try alternate names
    for ek, ev in excel_name_map.items():
        if ev == name and ek in raw_skills:
            return raw_skills[ek]
    return None

for name, base, cat, related in skills_def:
    excel = get_excel_data(name)
    
    # Build the object
    fields = [f"  {{ name: '{name}', base: {base}, cat: '{cat}', related: '{related}'"]
    
    if excel:
        desc = js_escape(excel['desc'])
        if desc:
            fields.append(f", desc: '{desc}'")
    
    # Check if there's a spec (for 科学)
    if name in spec_skills:
        sk = spec_skills[name]
        # Extract just the spec string (update base/cat/related if needed)
        fields.append(f", spec: {sk['spec']}")
    
    fields.append(" },")
    lines.append(' '.join(fields))

# Add spec-only skills (格斗, 射击, 技艺, 外语, 驾驶, 科学)
# These don't have desc from Excel but have spec
spec_order = ["格斗①", "格斗②", "格斗③", 
              "射击①", "射击②", "射击③",
              "技艺①", "技艺②", "技艺③",
              "外语①", "外语②", "外语③",
              "驾驶①"]

for name in spec_order:
    if name in spec_skills:
        sk = spec_skills[name]
        if any(d[0] == name for d in skills_def):
            continue  # Already added above
        fields = [f"  {{ name: '{name}', base: {sk['base']}, cat: '{sk['cat']}', related: '{sk['related']}'"]
        fields.append(f", spec: {sk['spec']}")
        fields.append(" },")
        lines.append(' '.join(fields))

lines.append('];')

output = '\n'.join(lines)

# Write preview file
with open('skills_data_preview.js', 'w', encoding='utf-8') as f:
    f.write('// Skills data preview — generated from Excel 附表\n')
    f.write('// Verify with: node --check skills_data_preview.js\n')
    f.write(output)
    f.write('\n')

print(f"\nGenerated skills_data_preview.js with {len(lines)} lines")
print(f"Skills with desc: {sum(1 for l in lines if 'desc:' in l)}")
print(f"Skills with spec: {sum(1 for l in lines if 'spec:' in l)}")

# Check for any raw newlines (should be none)
raw_newlines = [l for l in lines if '\n' in l and l.count("'") > 0]
if raw_newlines:
    print(f"WARNING: {len(raw_newlines)} lines have raw newlines!")
else:
    print("OK: No raw newlines in output")
