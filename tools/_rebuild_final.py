import openpyxl

wb = openpyxl.load_workbook('tools/COC7空白卡CY2lusFinal (1).xlsx', data_only=True)

# Get descriptions from 附表
ws = wb['附表']
desc_map = {}
for row_idx in range(95, 213):
    name_cell = ws.cell(row=row_idx, column=2).value
    desc_cell = ws.cell(row=row_idx, column=6).value
    if name_cell and str(name_cell).strip():
        name = str(name_cell).strip()
        desc = str(desc_cell).replace('\n', '\\n').replace("'", "\\'").strip() if desc_cell else ''
        desc_map[name] = desc

# All skills in order
skills = [
    ("会计",5,"investigate","INT"),
    ("人类学",1,"investigate","INT"),
    ("估价",5,"investigate","INT"),
    ("考古学",1,"investigate","INT"),
    ("侦查",25,"investigate","INT"),
    ("聆听",20,"investigate","INT"),
    ("追踪",10,"investigate","INT"),
    ("图书馆使用",20,"investigate","INT"),
    ("计算机使用",5,"investigate","INT"),
    ("信用评级",0,"investigate","INT"),
    ("克苏鲁神话",0,"investigate","INT"),
    ("锁匠",1,"investigate","DEX"),
    ("妙手",10,"investigate","DEX"),
    ("导航",10,"investigate","INT"),
    ("话术",5,"social","APP"),
    ("说服",10,"social","APP"),
    ("恐吓",15,"social","STR"),
    ("取悦",15,"social","APP"),
    ("乔装",5,"social","APP"),
    ("心理学",10,"social","INT"),
    ("格斗①",25,"combat","STR/DEX"),
    ("格斗②",25,"combat","STR/DEX"),
    ("格斗③",25,"combat","STR/DEX"),
    ("射击①",20,"combat","DEX"),
    ("射击②",20,"combat","DEX"),
    ("射击③",15,"combat","DEX"),
    ("投掷",20,"combat","DEX"),
    ("闪避",0,"combat","DEX/2"),
    ("攀爬",20,"special","STR"),
    ("跳跃",20,"special","STR"),
    ("游泳",20,"special","STR"),
    ("潜行",20,"special","DEX"),
    ("驾驶①",20,"special","DEX"),
    ("骑术",5,"special","DEX"),
    ("急救",30,"support","INT"),
    ("医学",1,"support","INT"),
    ("精神分析",1,"support","INT"),
    ("技艺①",5,"knowledge","INT"),
    ("技艺②",5,"knowledge","INT"),
    ("技艺③",5,"knowledge","INT"),
    ("电气维修",10,"knowledge","INT"),
    ("电子学",1,"knowledge","INT"),
    ("法律",5,"knowledge","INT"),
    ("历史",5,"knowledge","INT"),
    ("外语①",1,"knowledge","INT"),
    ("外语②",1,"knowledge","INT"),
    ("外语③",1,"knowledge","INT"),
    ("母语",0,"knowledge","EDU"),
    ("机械维修",10,"knowledge","INT"),
    ("博物学",10,"knowledge","INT"),
    ("神秘学",5,"knowledge","INT"),
    ("操作重型机械",1,"knowledge","INT"),
    ("科学",1,"knowledge","INT"),
    ("药学",1,"knowledge","INT"),
    ("催眠",1,"knowledge","INT"),
    ("读唇",1,"knowledge","INT"),
    ("爆破",1,"knowledge","INT"),
    ("潜水",1,"knowledge","INT"),
    ("动物驯养",5,"knowledge","INT"),
]

specs = {
    "格斗①": 'spec: {"type": "options", "defaultBase": 25, "options": [{"name": "斗殴", "base": 25}, {"name": "鞭子", "base": 5}, {"name": "电锯", "base": 10}, {"name": "链枷", "base": 10}, {"name": "绞具", "base": 15}, {"name": "斧", "base": 15}, {"name": "剑", "base": 20}, {"name": "矛", "base": 20}]}',
    "格斗②": 'spec: {"type": "options", "defaultBase": 25, "options": [{"name": "斗殴", "base": 25}, {"name": "鞭子", "base": 5}, {"name": "电锯", "base": 10}, {"name": "链枷", "base": 10}, {"name": "绞具", "base": 15}, {"name": "斧", "base": 15}, {"name": "剑", "base": 20}, {"name": "矛", "base": 20}]}',
    "格斗③": 'spec: {"type": "text", "defaultBase": 25}',
    "射击①": 'spec: {"type": "options", "defaultBase": 20, "options": [{"name": "手枪", "base": 20}, {"name": "步枪/霰弹枪", "base": 25}, {"name": "冲锋枪", "base": 15}, {"name": "机枪", "base": 10}, {"name": "弓术", "base": 15}, {"name": "喷射器", "base": 10}, {"name": "重武器", "base": 10}]}',
    "射击②": 'spec: {"type": "options", "defaultBase": 20, "options": [{"name": "手枪", "base": 20}, {"name": "步枪/霰弹枪", "base": 25}, {"name": "冲锋枪", "base": 15}, {"name": "机枪", "base": 10}, {"name": "弓术", "base": 15}, {"name": "喷射器", "base": 10}, {"name": "重武器", "base": 10}]}',
    "射击③": 'spec: {"type": "text", "defaultBase": 15}',
    "技艺①": 'spec: {"type": "options", "defaultBase": 5, "options": [{"name": "表演", "base": 5}, {"name": "美术", "base": 5}, {"name": "摄影", "base": 5}, {"name": "伪造", "base": 5}, {"name": "写作", "base": 5}, {"name": "书法", "base": 5}, {"name": "乐理", "base": 5}, {"name": "厨艺", "base": 5}, {"name": "裁缝", "base": 5}, {"name": "理发", "base": 5}, {"name": "建筑", "base": 5}, {"name": "舞蹈", "base": 5}, {"name": "酿酒", "base": 5}, {"name": "捕鱼", "base": 5}, {"name": "歌唱", "base": 5}, {"name": "制陶", "base": 5}, {"name": "雕塑", "base": 5}, {"name": "杂技", "base": 5}, {"name": "风水", "base": 5}, {"name": "技术制图", "base": 5}, {"name": "耕作", "base": 5}, {"name": "打字", "base": 5}, {"name": "速记", "base": 5}, {"name": "木匠", "base": 5}, {"name": "莫里斯舞蹈", "base": 5}, {"name": "歌剧歌唱", "base": 5}, {"name": "粉刷匠与油漆工", "base": 5}, {"name": "吹真空管", "base": 5}]}',
    "技艺②": 'spec: {"type": "options", "defaultBase": 5, "options": [{"name": "表演", "base": 5}, {"name": "美术", "base": 5}, {"name": "摄影", "base": 5}, {"name": "伪造", "base": 5}, {"name": "写作", "base": 5}, {"name": "书法", "base": 5}, {"name": "乐理", "base": 5}, {"name": "厨艺", "base": 5}, {"name": "裁缝", "base": 5}, {"name": "理发", "base": 5}, {"name": "建筑", "base": 5}, {"name": "舞蹈", "base": 5}, {"name": "酿酒", "base": 5}, {"name": "捕鱼", "base": 5}, {"name": "歌唱", "base": 5}, {"name": "制陶", "base": 5}, {"name": "雕塑", "base": 5}, {"name": "杂技", "base": 5}, {"name": "风水", "base": 5}, {"name": "技术制图", "base": 5}, {"name": "耕作", "base": 5}, {"name": "打字", "base": 5}, {"name": "速记", "base": 5}, {"name": "木匠", "base": 5}, {"name": "莫里斯舞蹈", "base": 5}, {"name": "歌剧歌唱", "base": 5}, {"name": "粉刷匠与油漆工", "base": 5}, {"name": "吹真空管", "base": 5}]}',
    "技艺③": 'spec: {"type": "text", "defaultBase": 5}',
    "外语①": 'spec: {"type": "text", "defaultBase": 1}',
    "外语②": 'spec: {"type": "text", "defaultBase": 1}',
    "外语③": 'spec: {"type": "text", "defaultBase": 1}',
    "驾驶①": 'spec: {"type": "text", "defaultBase": 20}',
    "科学": 'spec: {"type": "options", "defaultBase": 1, "options": [{"name": "数学", "base": 10}, {"name": "地质学", "base": 1}, {"name": "化学", "base": 1}, {"name": "生物学", "base": 1}, {"name": "物理学", "base": 1}, {"name": "天文学", "base": 1}, {"name": "气象学", "base": 1}, {"name": "药学", "base": 1}, {"name": "工程学", "base": 1}, {"name": "密码学", "base": 1}, {"name": "制图学", "base": 1}, {"name": "人类学", "base": 1}, {"name": "心理学", "base": 1}]}',
}

lines = ["const ALL_SKILLS = ["]
for name, base, cat, related in skills:
    parts = [f"  {{ name: '{name}', base: {base}, cat: '{cat}', related: '{related}'"]
    desc = desc_map.get(name, '')
    if desc:
        parts.append(f", desc: '{desc}'")
    if name in specs:
        parts.append(f", {specs[name]}")
    parts.append(" },")
    lines.append(' '.join(parts))
lines.append("];")

new_block = '\n'.join(lines)

# Replace in file
with open('pages/coc7-gen/coc7-gen.js', 'r', encoding='utf-8') as f:
    content = f.read()

start = content.find('const ALL_SKILLS = [')
# Find the matching ];
brace_count = 0
end = start
for i in range(start, len(content)):
    if content[i] == '[':
        brace_count += 1
    elif content[i] == ']':
        brace_count -= 1
        if brace_count == 0:
            end = i + 1  # include ]
            break
# Also include any trailing whitespace/newline/semicolon
while end < len(content) and content[end] in ' \t\n\r;':
    end += 1

new_content = content[:start] + new_block + '\n' + content[end:]
with open('pages/coc7-gen/coc7-gen.js', 'w', encoding='utf-8') as f:
    f.write(new_content)

print(f'Done: {len(skills)} skills, {len(lines)} lines')
