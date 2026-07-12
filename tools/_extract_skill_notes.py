# -*- coding: utf-8 -*-
"""Extract skill click notes from 人物卡 sheet and 附表 sheet, then generate clean desc data."""
import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('COC7空白卡CY2lusFinal (1).xlsx', data_only=True)

# Sheet indices: 2=人物卡, 7=附表, 8=技能注释
ws_char = wb.worksheets[2]  # 人物卡
ws_fubiao = wb.worksheets[7]  # 附表
ws_notes = wb.worksheets[8]  # 技能注释

# === Part 1: Extract skill names and notes from 附表 (the main skill descriptions) ===
print("=" * 60)
print("PART 1: 附表 skill data (rows 95-213, col B=name, col F=desc)")
print("=" * 60)
fubiao_skills = {}
for r in range(95, 213):
    name = ws_fubiao.cell(row=r, column=2).value
    desc = ws_fubiao.cell(row=r, column=6).value
    usage = ws_fubiao.cell(row=r, column=5).value  # 特殊适用
    if name and str(name).strip():
        n = str(name).strip()
        d = str(desc).strip() if desc else ''
        u = str(usage).strip() if usage else ''
        fubiao_skills[n] = {'desc': d, 'usage': u}
        if d:
            print(f"  [{n}] desc={d[:80]}...")
        else:
            print(f"  [{n}] (no desc)")

print(f"\nTotal skills in 附表: {len(fubiao_skills)}")

# === Part 2: Extract skill click notes from 人物卡 ===
# The 人物卡 sheet has skills in rows, with notes (comments) embedded
print("\n" + "=" * 60)
print("PART 2: 人物卡 - looking for skill notes/comments")
print("=" * 60)

# Look at columns around skills area (typically row 16+)
for r in range(14, 100):
    row_data = {}
    for c in [2, 3, 4, 5, 6, 7, 8, 9, 10, 60, 61, 62, 70, 80]:
        v = ws_char.cell(row=r, column=c).value
        if v and str(v).strip():
            row_data[c] = str(v).strip()[:100]
    if row_data:
        print(f"  Row {r}: {row_data}")

# === Part 3: Look at 技能注释 sheet ===
print("\n" + "=" * 60)
print("PART 3: 技能注释 sheet")
print("=" * 60)
for r in range(1, min(ws_notes.max_row + 1, 50)):
    row_data = {}
    for c in range(1, 15):
        v = ws_notes.cell(row=r, column=c).value
        if v and str(v).strip():
            row_data[c] = str(v).strip()[:120]
    if row_data:
        print(f"  Row {r}: {json.dumps(row_data, ensure_ascii=False)}")

# === Part 4: Check for comments/notes in cells ===
print("\n" + "=" * 60)
print("PART 4: Cell comments in 人物卡 (first 100 rows)")
print("=" * 60)
for r in range(1, 101):
    for c in range(1, 90):
        cell = ws_char.cell(row=r, column=c)
        if cell.comment:
            print(f"  Row {r} Col {c}: comment='{cell.comment.text[:100]}'")
