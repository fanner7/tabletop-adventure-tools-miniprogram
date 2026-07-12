import openpyxl, json, re

wb = openpyxl.load_workbook('tools/COC7空白卡CY2lusFinal (1).xlsx', data_only=True)

# Get descriptions from 附表
ws = wb['附表']
desc_map = {}
for row_idx in range(95, 213):
    name_cell = ws.cell(row=row_idx, column=2).value
    desc_cell = ws.cell(row=row_idx, column=6).value
    usage_cell = ws.cell(row=row_idx, column=5).value
    if name_cell and str(name_cell).strip():
        name = str(name_cell).strip()
        desc = str(desc_cell).replace('\n', '\\n').replace("'", "\\'").strip() if desc_cell else ''
        usage = str(usage_cell).replace('\n', '\\n').replace("'", "\\'").strip() if usage_cell else ''
        desc_map[name] = {'desc': desc, 'usage': usage}

# Manual skill definitions (known-good structure, matching Excel names)
# We preserve the spec objects exactly as they were
skills_def = [
    ("会计", 5, "investigate", "INT"),
    ("人类学", 1, "investigate", "INT"),
    ("估价", 5, "investigate", "INT"),
    ("考古学", 1, "investigate", "INT"),
    ("侦查", 25, "investigate", "INT"),
    ("聆听", 20, "investigate", "INT"),
    ("追踪", 10, "investigate", "INT"),
    ("图书馆使用", 20, "investigate", "INT"),
    ("计算机使用", 5, "investigate", "INT"),
    ("信用评级", 0, "investigate", "INT"),
    ("克苏鲁神话", 0, "investigate", "INT"),
    ("锁匠", 1, "investigate", "DEX"),
    ("妙手", 10, "investigate", "DEX"),
    ("导航", 10, "investigate", "INT"),
    ("话术", 5, "social", "APP"),
    ("说服", 10, "social", "APP"),
    ("恐吓", 15, "social", "STR"),
    ("取悦", 15, "social", "APP"),
    ("乔装", 5, "social", "APP"),
    ("心理学", 10, "social", "INT"),
    ("闪避", 0, "combat", "DEX/2"),
    ("投掷", 20, "combat", "DEX"),
    ("攀爬", 20, "special", "STR"),
    ("跳跃", 20, "special", "STR"),
    ("游泳", 20, "special", "STR"),
    ("潜行", 20, "special", "DEX"),
    ("骑术", 5, "special", "DEX"),
    ("急救", 30, "support", "INT"),
    ("医学", 1, "support", "INT"),
    ("精神分析", 1, "support", "INT"),
    ("电气维修", 10, "knowledge", "INT"),
    ("电子学", 1, "knowledge", "INT"),
    ("法律", 5, "knowledge", "INT"),
    ("历史", 5, "knowledge", "INT"),
    ("母语", 0, "knowledge", "EDU"),
    ("机械维修", 10, "knowledge", "INT"),
    ("博物学", 10, "knowledge", "INT"),
    ("神秘学", 5, "knowledge", "INT"),
    ("操作重型机械", 1, "knowledge", "INT"),
    ("药学", 1, "knowledge", "INT"),
    ("催眠", 1, "knowledge", "INT"),
    ("读唇", 1, "knowledge", "INT"),
    ("爆破", 1, "knowledge", "INT"),
    ("潜水", 1, "knowledge", "INT"),
    ("动物驯养", 5, "knowledge", "INT"),
]

# Generate lines
lines = ["const ALL_SKILLS = ["]

spec_map = {
    "格斗①": 'spec: {"type": "options", "defaultBase": 25, "options": [{"name": "斗殴", "base": 25}, {"name": "鞭子", "base": 5}, {"name": "电锯", "base": 10}, {"name": "链枷", "base": 10}, {"name": "绞具", "base": 15}, {"name": "斧", "base": 15}, {"name": "剑", "base": 20}, {"name": "矛", "base": 20}]}',
    "格斗②": 'spec: {"type": "options", "defaultBase": 25, "options": [{"name": "斗殴", "base": 25}, {"name": "鞭子", "base": 5}, {"name": "电锯", "base": 10}, {"name": "链枷", "base": 10}, {"name": "绞具", "base": 15}, {"name": "斧", "base": 15}, {"name": "剑", "base": 20}, {"name": "矛", "base": 20}]}',
    "格斗③": 'spec: {"type": "text", "defaultBase": 25}',
    "射击①": 'spec: {"type": "options", "defaultBase": 20, "options": [{"name": "手枪", "base": 20}, {"name": "步枪/霰弹枪", "base": 25}, {"name": "冲锋枪", "base": 15}, {"name": "机枪", "base": 10}, {"name": "弓术", "base": 15}, {"name": "喷射器", "base": 10}, {"name": "重武器", "base": 10}]}',
    "射击②": 'spec: {"type": "options", "defaultBase": 20, "options": [{"name": "手枪", "base": 20}, {"name": "步枪/霰弹枪", "base": 25}, {"name": "冲锋枪", "base": 15}, {"name": "机枪", "base": 10}, {"name": "弓术", "base": 15}, {"name": "喷射器", "base": 10}, {"name": "重武器", "base": 10}]}',
    "射击③": 'spec: {"type": "text", "defaultBase": 15}',
    "技艺①": 'spec: {"type": "options", "defaultBase": 5, "options": [{"name": "表演", "base": 5}, {"name": "美术", "base": 5}, {"name": "摄影", "base": 5}, {"name": "伪造", "base": 5}, {"name": "写作", "base": 5}, {"name": "书法", "base": 5}, {"name": "乐理", "base": 5}, {"name": "厨艺", "base": 5}, {"name": "裁缝", "base": 5}, {"name": "理发", "base": 5}, {"name": "建筑", "base": 5}, {"name": "舞蹈", "base": 5}, {"name": "酿酒", "base": 5}, {"name": "捕鱼", "base": 5}, {"name": "歌唱", "base": 5}, {"name": "制陶", "base": 5}, {"name": "雕塑", "base": 5}, {"name": "杂技", "base": 5}, {"name": "风水", "base": 5}, {"name": "技术制图", "base": 5}, {"name": "耕作", "base": 5}, {"name": "打字", "base": 5}, {"name": "速记", "base": 5}, {"name": "木匠", "base": 5}, {"name": "莫里斯舞蹈", "base": 5}, {"name": "歌剧歌唱", "base": 5}, {"name": "粉刷匠与油漆工", "base": 5}, {"name": "吹真空管", "base": 5}]}',
    "技艺②": 'spec: {"type": "options", "defaultBase": 5, "options": [{"name": "表演", "base": 5}, {"name": "美术", "base": 5}, {"name": "摄影", "base": 5}, {"name": "伪造", "base": 5}, {"name": "写作", "base": 5}, {"name": "书法", "base": 5}, {"name": "乐理", "base": 5}, {"name": "厨艺", "base": 5}, {"name": "裁缝", "base": 5}, {"name": "理发", "base": 5}, {"name": "建筑", "base": 5}, {"name": "舞蹈", "base": 5}, {"name": "酿酒", "base": 5}, {"name": "捕鱼", "base": 5}, {"name": "歌唱", "base": 5}, {"name": "制陶", "base": 5}, {"name": "雕塑", "base": 5}, {"name": "杂技", "base": 5}, {"name": "风水", "base": 5}, {"name": "技术制图", "base": 5}, {"name": "耕作", "base": 5}, {"name": "打字", "base": 5}, {"name": "速记", "base": 5}, {"name": "木匠", "base": 5}, {"name": "莫里斯舞蹈", "base": 5}, {"name": "歌剧歌唱", "base": 5}, {"name": "粉刷匠与油漆工", "base": 5}, {"name": "吹真空管", "base": 5}]}',
    "技艺③": 'spec: {"type": "text", "defaultBase": 5}',
    "外语①": 'spec: {"type": "text", "defaultBase": 1}',
    "外语②": 'spec: {"type": "text", "defaultBase": 1}',
    "外语③": 'spec: {"type": "text", "defaultBase": 1}',
    "驾驶①": 'spec: {"type": "text", "defaultBase": 20}',
    "科学": 'spec: {"type": "options", "defaultBase": 1, "options": [{"name": "数学", "base": 10}, {"name": "地质学", "base": 1}, {"name": "化学", "base": 1}, {"name": "生物学", "base": 1}, {"name": "物理学", "base": 1}, {"name": "天文学", "base": 1}, {"name": "气象学", "base": 1}, {"name": "药学", "base": 1}, {"name": "工程学", "base": 1}, {"name": "密码学", "base": 1}, {"name": "制图学", "base": 1}, {"name": "人类学", "base": 1}, {"name": "心理学", "base": 1}]}',
}

for name, base, cat, related in skills_def:
    info = desc_map.get(name, {'desc': '', 'usage': ''})
    desc = info['desc']
    usage = info['usage']
    
    parts = [f"  {{ name: '{name}', base: {base}, cat: '{cat}', related: '{related}'"]
    if usage:
        parts.append(f", usage: '{usage}'")
    
    # Check if it has a spec
    if name in spec_map:
        # For spec skills, the desc might come from spec definitions
        # Use desc from spec if available
        pass
    
    if desc:
        parts.append(f", desc: '{desc}'")
    
    if name in spec_map:
        parts.append(f", {spec_map[name]}")

    parts.append(" },")
    lines.append(' '.join(parts))

lines.append("];")

new_skills_block = '\n'.join(lines)

# Now find and replace ALL_SKILLS in the current file
with open('pages/coc7-gen/coc7-gen.js', 'r', encoding='utf-8') as f:
    content = f.read()

# Find the bounds
start = content.find('const ALL_SKILLS = [')
end = content.find('\n];', start)
if end == -1:
    end = content.find('];', start)
# Find the actual end of ALL_SKILLS
end = content.find('];', start)
if end >= 0:
    end += 2  # include ];
else:
    print("ERROR: cannot find end of ALL_SKILLS")
    exit(1)

new_content = content[:start] + new_skills_block + content[end:]

with open('pages/coc7-gen/coc7-gen.js', 'w', encoding='utf-8') as f:
    f.write(new_content)

print(f'ALL_SKILLS rebuilt: {len(skills_def)} base + {len(spec_map)} spec skills')
print(f'Total lines in ALL_SKILLS: {len(lines)}')
