# -*- coding: utf-8 -*-
"""Look at 人物卡 skill rows and their surrounding columns for notes/comments/remarks."""
import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('COC7空白卡CY2lusFinal (1).xlsx', data_only=True)
ws = wb.worksheets[2]  # 人物卡

# Look at rows 16-49 (skill area) across ALL columns
for r in range(16, 50):
    skill_name = ws.cell(row=r, column=6).value
    if not skill_name or not str(skill_name).strip():
        continue
    name = str(skill_name).strip()
    
    # Check all columns for any non-None data
    extras = {}
    for c in range(1, 90):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip():
            extras[c] = str(v).strip()[:150]
    
    # Also check for notes/comments
    comments = {}
    for c in range(1, 90):
        cell = ws.cell(row=r, column=c)
        if cell.comment:
            comments[c] = cell.comment.text[:200]
    
    if len(extras) > 2 or comments:  # More than just the standard cols
        print(f"\n=== Row {r}: {name} ===")
        if len(extras) > 4:
            print(f"  All cols: {json.dumps(extras, ensure_ascii=False)}")
        else:
            for c, v in extras.items():
                print(f"  col {c}: {v}")
        if comments:
            for c, v in comments.items():
                print(f"  COMMENT col {c}: {v}")
