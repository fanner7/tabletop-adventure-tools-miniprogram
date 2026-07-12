import openpyxl, json, re

wb = openpyxl.load_workbook('tools/COC7空白卡CY2lusFinal (1).xlsx', data_only=True)

# Read 附表 for skill descriptions + 人物卡 for skill names
# 附表 B:K from row 95
ws = wb['附表']
skills_data = {}
for row_idx in range(95, 213):
    b = ws.cell(row=row_idx, column=2).value  # name
    f = ws.cell(row=row_idx, column=6).value  # desc
    if b and str(b).strip():
        name = str(b).strip()
        desc = str(f).strip() if f else ''
        skills_data[name] = desc

# Now map these to the current ALL_SKILLS structure
# Read the current JS file to get the skill definitions
with open('pages/coc7-gen/coc7-gen.js', 'r', encoding='utf-8') as f:
    content = f.read()

# Get skill definitions from current file
# Find ALL_SKILLS
start = content.find('const ALL_SKILLS = [\n  {')
end = content.find('\n];\n\nconst OCCUPATIONS')
if end == -1:
    end = content.find('\n];\n\n// ----------', start)

skills_block = content[start:end]

# Parse individual skill entries
# Each line is: { name: 'X', base: N, cat: 'X', ... },
skills = []
for line in skills_block.split('\n'):
    line = line.strip()
    if line.startswith('{ name:') and line.endswith('},'):
        skills.append(line)

# Print what we found
for s in skills[:5]:
    print(s[:80])
print(f'... total {len(skills)} skills')
print(f'Excel descriptions for {len(skills_data)} skills')

# Now build the new ALL_SKILLS
lines_out = ['const ALL_SKILLS = [']
for s in skills:
    # Extract the name
    m = re.search(r"name: '([^']+)'", s)
    if m:
        name = m.group(1)
        # Replace desc if we have it from Excel
        if name in skills_data:
            excel_desc = skills_data[name].replace('\n', '\\n').replace("'", "\\'")
            # Replace desc: '...' in the skill definition
            s = re.sub(r"desc: '[^']*'", f"desc: '{excel_desc}'", s)
            # If there's no desc yet, add it
            if "desc: '" not in s and s.endswith("},"):
                s = s[:-2] + f", desc: '{excel_desc}' },"
    lines_out.append('  ' + s)

lines_out.append('];')
new_skills = '\n'.join(lines_out)

# Replace ALL_SKILLS in the content
new_content = content[:start] + new_skills + content[end:]

with open('pages/coc7-gen/coc7-gen.js', 'w', encoding='utf-8') as f:
    f.write(new_content)

print(f'Wrote {len(lines_out)} lines for ALL_SKILLS')
